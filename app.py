import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, HRFlowable, PageBreak)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

st.set_page_config(page_title="Yellow Sheet", page_icon="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAAXHUlEQVR4nO3de6yfdX3A8e9pTwtSpNBSoV6GUha5CDh3oiIYL8HLcpx3O51ZujBNluGWMRf3x+IymZvZ1JXpiMvi3OoyszUxmY66G7LpjKAc3MAhBS0ORbDUgxaspaXt2R9wyvndb8/vuXw/r1dilF/P4Ty/Xzzn834+z3N+nUn0tWFu61LVxwDAeB5c2DlT9THUlRcmGfIAEUWPg5BP3sAHoF20IAjxZA18AEaVexBk++QMfQCKkmMMZPWEDH0Api2XGMjiSRj8AJSt6SHQ6IM3+AGoWlNDoJEHbfADUDdNC4FGHazBD0DdNSUEGnGQBj8ATVP3EKj1wRn8ADRdXUNgVdUH0IvhD0AO6jrPalcldX2hAGBSddoG1GoDYPgDkLM6zblalEidXhAAKEPV24DKNwCGPwARVT3/Kg2Aqp88AFSpyjlYyfrB4AeAVmVfEih9A2D4A0CnsudjqQFg+ANAb2XOydICwPAHgMHKmpelBIDhDwDDK2NuTj0ADH8AGN205+dUA8DwB4DxTXOOTi0ADH8AmNy05ulUAsDwB4DiTGOuFh4Ahj8AFK/o+VpoABj+ADA9Rc7ZwgLA8AeA6Stq3lb+twECAOUrJACc/QNAeYqYuxMHgOEPAOWbdP5OFACGPwBUZ5I57B4AAAho7ABw9g8A1Rt3Ho8VAIY/ANTHOHPZJQAACGjkAHD2DwD1M+p8HikADH8AqK9R5rRLAAAQ0NAB4OwfAOpv2HltAwAAAQ0VAM7+AaA5hpnbNgAAENDAAHD2DwDNM2h+2wAAQEB9A8DZPwA0V785bgMAAAH1DABn/wDQfL3muQ0AAATUNQCc/QNAPrrNdRsAAAhIAABAQB0BYP0PAPlpn+82AAAQkAAAgIBaAsD6HwDytXLO2wAAQEACAAACEgAAENDxAHD9HwDytzzvbQBGtLh9Z9WHAAATEwBjEAEANJ0AGMHKwS8CAGiyVSm5/j+MxctP6XxMBADQQBvmti7ZAAzh+PDf9Y7OPxMBADSQABig48xfBACQAQHQw+Llp3Rd+6eURAAAjScAuug5+FcSAQA0mABoM9TwXyYCAGgoAdBm4/UPjfYJIgCABlrlVwA7iQAAcmcD0IMIACBnAqAPEQBArgTAIPMfH+3ju0QAANSNABjGBBGw8aqtBR8MAExOAAzLJgCAjAiAPjqu348SAaMGAwCUSACMapjB/vjHWP8DUFezVR9AI81/vPuKvwFn/Ys3XVv1IUCxbtwktmEMNgA9DPz1vfZhX+Phv3jTtcf/Azny67YwOgEwieWhX9Phb+gTiQiA0QiASfUY/lWvJA1+IhIBMDwB0EXTf4gY/kTW9O9fKIsAyIzhDyIAhiEApqCq9b/hD08QAdCfAMiE4Q+dRAD0JgDaNPEHhuEPvTXxexrKIAAKVvb63/CHwUQAdBIAQAgiAFoJgBWa9gPC2T+Mpmnf4zBNAgAIRQTAYwRAgap+9z9gOCIABMBxTfuBYP0Pk2na9zwUTQAAYYkAIhMABbH+h2YSAUQlAIDwRAARCYDkmx/wc4B4BEABrP8hDyKASAQAwAoigCjCB4BvdqCdnwtEED4AALoRAeROAEzI9X/IlwggZ6EDwDc3MIifE+QqdAAADEMEkCMBMAHrf4hDBJAbAQAwJBFATmarPoCqhPtGvnFT1UcAWZj0Z4fNIXVhAzAm38QANJkAAICAQgZAuPU/ALQJGQAAEJ0AGIPr/wA0XbgAsP4HgIABAAAIgJFZ/wOQAwEAAAGFCgDX/wHgMaECYFLW/wDkIuzfBcATDhyaTZf+8avTd394UsvjH3zzLemKS/f0/Lyrds6lT954dstjl53zQPrHX/vPNDPzxGMPHlibPvXVZ6XP7z4z7b5/fdp/cG1KaSmdetKjaf2TDqcz1x9M5535UDp/84/Si7bsS2dv+nGRTw+ALsIEgPV/b+tOOJKu+YWb05v+4iUtj1993UXp555zX9q8/mDH53x5z6b0tze1Dv8nrT2a/uytN7cM/3/7xuZ05adekB48sLbj37H3odVp70Mnprv2npK+eNcZKaWUfuWyb6U/edPXCnhWAPQT5xLAhddVfQS19tJn701vf8G3Wx57+JE16T2ffl7Hxx46sipdtXMuLS21Pv7e+dvSMzceOP7Pt373tLTtE5d2Hf4AVCvMBiCl9FgEfP01Y31qhOv/73/9/6Qbdp+Z7t//pOOPfe7rT0vX3fb09JqL7j3+2If//fz0rQee3PK5z3/WD9I7X/ytlsf+8HMXpsNHn2jMk9YeSe977a3pVRfcn04/+ZF04NBsunPv+nT9HWemv7/5men7K74uANMVZgOw8fJPPvY/bAJ6OuXER9OH33JLx+Pv+fTz0kOPrEkppXTH/evTRz5/bsufnzB7NH30rTenVTNPrAQOH12V/uubT2n5uHe/8hvpikv3pKed+pN0wuyxtGHd4XTJ2fvSe+e/nm79vevSn25dSGc8+ZEpPDOoD5cjqYswAdBCBPT0qgvuS29+3ndaHtv70Inp9z97cTq2NJN+8x/m0qNHW/9v8zuvvj2d85SHWx77wcMntJz9p5TS00/7Sc+vO7tqKW275O707ld+Y8JnAPUnAqiDmAGQ0kgREGH9v9IH3vi1dPrJh1oe++RNZ6crP/X8tHDPxpbHn/uMB9O7XnZnx79j7eyxjsf+8os/nfY9fGKxBwsNJQKoWqgAOH4ZYJlNQFcb1h1OH3xz66WApaWUdi6c1fLYmtXH0kffdnNavartbsCU0uknH+r47YFb7tmYLr56Pv3ixy9LH7nh3PSFu85I+w+uKf4JQEOIAKoUKgC6EgFdvfbie9PPX3xv34/5rVfckc7fvL/nn7/zxd/seOzQkdXpX29/anrfP12U3vixl6Rzfvf16WUfekW65vrz0g9/4rcFiEcEUBUBkFLfCIi2/l/pg2/6WjrtpMNd/+z8zfvTVZff0ffzf/3ld6atc/f0/ZhjSzPptu+dlv5g14XpZ98/n/7l9qeOfbzQVCKAKgiAZTYBHTY9+ZH0R2/4747HV69aSn/+tq+mNas7r/OvtGpmKX3s7V9Jf/eOL6VLt+xreYOgbvYfXJOu+JsXpTu/f8okhw2NJAIoW7gA6LgPYCUR0OF1z+28DPBTGw6ki5/xw6H/Ha++4L702Xf9R9p99WfSX//yl9OVL70zPf9ZP0izXe4dOHRkVfqrL50z0TFDU4kAyhQuAAYSAVNz+smH0msvvjdd/bpb0z//xg3pjqs/k37phXd3fNxt3zutgqODehABlEUAdPN4BES+/l+GDesOpw+95ZZ04pqjLY8/enTAtQLInAigDCEDoO9lgGU2ARN7w8dekm7YfWbH3xmw0sHDs+lI2xsGdfvLhyAaEcC0xfq7AEYwVCTQ11fu3pS+eNcZ6ayNB9IbfuY76UVb9qULNu9PG04+lA4eXp3+975T0wc+95x05FjrGf/Lnr23oiOGelncvtMmkqkRAF0Y/sW6Z3Fduub689I115838GPP2nig428lhKxcsm/ET7hyKocBIS8BpGTIl2HdCUdG+vjzNu9Pn/7VL3TcEwBA8WwA2giD4tz+vs+mr3779PSVu09Pt957Wvq/xXXp/v0npR8/MpuOLs2kdWuPpM2nHkzPeeqP0vyF30vzF93b9VcDASieAFjB8O90wuzRsW9GWrv6WLrsnAfSZec8UPBRATCpsJcA2hn+AEQSOgAMfQCiCh0Ay4QAANGEDwDDH4CIwgcAAEQkAAAgIAEAAAF5H4AoRn77UQByZgMAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACGi26gOgXvYs7NhV9TFAE22Z2zZf9THAKGwAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABeStgWng7U4AYbAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQLNVHwD1smdhx66qjwG62TK3bb7qY4Cc2AAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAALyVsC08HarADHYAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICAZqs+AOplz8KOXVUfAzTRlrlt81UfA4zCBgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJC3AqaFtzMFiMEGAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQ0W/UBUC97FnbsqvoYoJstc9vmqz4GyIkNAAAEJAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIG8FTAtvtwoQgw0AAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACGi26gOgXvYs7NhV9TFAE22Z2zZf9THAKGwAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABeStgWng7U4AYbAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQLNVHwD1smdhx66qjwG62TK3bb7qY4Cc2AAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAALyVsC08HarADEIAICy3Lhp5E9Z3P7Yf2+8amvBB0N0AgBgGsYY9v0sbt/Z8ZgoYBICAKAoBQ/9QVZGgRhgVAIAYBIlD/1exACjEgAAo5pk6K8/d/TP2b97pA8XAwxDAAAMa9TBP86wH/bfM2QULMeAEKCdAAAYZNjBX9TAH+drDQgCIUA7bwQE0M8ww3/9ueUO/wmOodtvExCTDQBAN4MGf9UDv5eVx9VjK2AbQEo2AACd+g3/OpztD2vAsdoGxGYDALBs0OBvquVj77IRsA2IywYAIKV8h/9KtgGsIAAAeg3/Jq37h9XnOYmAWAQAEFu/4Z8zERCeewCAmCKs/AdZf677AgKzAQDiibTyH8QlgbAEABBL1JX/ICIgHAEAEH34L/M6hCIAgDi6nf0beq26vB62AHkSAEAMhv/wREAIfguAFnsWduyq+higaFse/e35jgcN//66/IbA4vadfjMgIzYAQNYM/wnYBGRNAABAQAIAyJaz/wLYAmRLAABZMvwLJAKyJACAGAz/yXj9siMAgOx0PfuncLYAzSYAgKxY/U+RSwFZEQBA3gz/Ynk9syEAgGxY/VfDFqCZBACQL2er0+F1zcLMhrmtS1UfBKNbvOnaqg8B6sV7/Zer7W2CU0reJrhhbACAPAUf/huvuGi6XyD465sDAQCQmeXhP/UIoNEEANB83db/QbUP/TIjwM2AzSIAgPxYT7eYWgR4nRtNAABkot+gdzmAdgIAaLb29X/Qs9JBA37xE7dN5wu3vd4uAzSHAABouMqGP40mAAAazPBnXAIAaK7gd//X9bq+ywDNIACAfAS9/t9LaWf/XvdGEgAADWT1z6QEAEDDGP4UQQAANIjhT1EEANBMAW8AbNLwdyNg/QkAIA+Z34hW1zv+j8v89c+RAADIQJ3O/mkGAQBQc01a/dMcAgCgxgx/pkUAANSU4c80CQCAGjL8mTYBADCGad6VX/s7/smCAAAY0fKArmpQO/unCAIAYATtQ7/oCLD6pywCAGBIvYZzURFg+FMmAQAwhGmv+w1/yiYAAIYwaABPEgiGP1UQAABDmkYEuOOfqggAIA/7d5fyZaa5CRjn69VGSa8/xREAQDNdsq+yL11UBOS8+t941daqD4EBBADAFAwa7jkPf5pBAACMYZgBPe6vDRr+lEEAAIxpnAgw/KkLAQDko4Ib0UYZ2Nne8e8GwEYSAEBzVXgj4EpF3RSYy9m/GwCbQQAAFGDSCMhl+NMcAgCgIOMOccOfKggAoNnaLwNUfD161GHe+OHf9npb/zeHAGiojS+8supDACbU+OFPowkAID8N2wI0lrv/G00AAM1Xk98GWKk9Agb9cw6s/5tFADSYywBQb8tDvtd/Q5UEAJCnmqynsz3zr8nry/gEQMPZAsDjangZIBLr/+YRABkQAdCDs9Tp8LpmQQBkQgRAsgWoiLP/ZhIAGREB0IWz1WJ5PbMhADIjAgiv2xbA0CpGl9fR2X9zCYAMiQDCcymgFIZ/swmATIkAaGMLMBmvX3ZmNsxtXar6IJiuxZuurfoQoBo3bup8bP255R9H01n9Z0kABCMGCEcETMbwz9Zs1QdAuVwaIJrF7VUfAdSTewCArHU9W3U9ezjO/rMmAIDsiYAxGP7ZEwBACCJgBIZ/CAIACEMEDMHwD0MAAIiAx3gdQhEAQCg9z2ajD78ez9/Zf74EABBO3wiIFgJ9nrPhnzdvBASEtrh9Z/c/iPBmQQZ/aDYAQGhhLwkY/uEJACC8UJcErPx5nAAASAOGXy4R0Od5GP7xuAcAoE3P+wJSaua9AQY/XdgAALQZuA1oykZgwLEa/rHZAAD00XcbsKxOW4Eh4sTgJyUbAIC+hhqWddgKDHkMhj/LbAAAhjTUNmClaW4GRgwOg592AgBgRCOHwErjRMEE2wWDn14EAMAEJoqBKTH0GYYAAChIlTFg6DOqmZRSEgEAxSojBgx9JiEAAEoySRQY9hRNAABAQN4HAAACEgAAEJAAAICABAAABLQqpZQeXNg5U/WBAADleHBh54wNAAAEJAAAICABAAABHQ8A9wEAQP6W570NAAAEJAAAICABAAABtQSA+wAAIF8r57wNAAAEJAAAIKCOAHAZAADy0z7fbQAAICABAAABdQ0AlwEAIB/d5roNAAAE1DMAbAEAoPl6zXMbAAAIqG8A2AIAQHP1m+M2AAAQ0MAAsAUAgOYZNL9tAAAgoKECwBYAAJpjmLltAwAAAQ0dALYAAFB/w85rGwAACGikALAFAID6GmVOj7wBEAEAUD+jzmeXAAAgoLECwBYAAOpjnLk89gZABABA9cadxy4BAEBAEwWALQAAVGeSOTzxBkAEAED5Jp2/hVwCEAEAUJ4i5q57AAAgoMICwBYAAKavqHlb6AZABADA9BQ5Zwu/BCACAKB4Rc/XqdwDIAIAoDjTmKtTuwlQBADA5KY1T6f6WwAiAADGN805OvVfAxQBADC6ac/PUt4HQAQAwPDKmJulvRGQCACAwcqal6W+E6AIAIDeypyTpb8VsAgAgE5lz8dKh/GGua1LVX59AKhaVSfGlf5lQLYBAERW5Rys/G8DFAEARFT1/KvV8HVJAIDcVT34l1W+AVipLi8KAExDneZcbQ6knW0AALmo0+BfVqsNwEp1fLEAYFR1nWe1PKh2tgEANE1dB/+yWh9cOyEAQN3VffAva8RBthMCANRNUwb/skYdbDshAEDVmjb4lzXyoNsJAQDK1tTBv6zRB99OCAAwbU0f/MuyeBLdiAEAipLL0F8puyfUjRgAYFQ5Dv2Vsn5yvQgCANrlPvDbhXqyvQgCgHiiDfx2/w8eWPS/OI7SkAAAAABJRU5ErkJggg==", layout="wide")
# ── App logo + title ──────────────────────────────────────────────────────────
LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAgAAAAIACAYAAAD0eNT6AAAXHUlEQVR4nO3de6yfdX3A8e9pTwtSpNBSoV6GUha5CDh3oiIYL8HLcpx3O51ZujBNluGWMRf3x+IymZvZ1JXpiMvi3OoyszUxmY66G7LpjKAc3MAhBS0ORbDUgxaspaXt2R9wyvndb8/vuXw/r1dilF/P4Ty/Xzzn834+z3N+nUn0tWFu61LVxwDAeB5c2DlT9THUlRcmGfIAEUWPg5BP3sAHoF20IAjxZA18AEaVexBk++QMfQCKkmMMZPWEDH0Api2XGMjiSRj8AJSt6SHQ6IM3+AGoWlNDoJEHbfADUDdNC4FGHazBD0DdNSUEGnGQBj8ATVP3EKj1wRn8ADRdXUNgVdUH0IvhD0AO6jrPalcldX2hAGBSddoG1GoDYPgDkLM6zblalEidXhAAKEPV24DKNwCGPwARVT3/Kg2Aqp88AFSpyjlYyfrB4AeAVmVfEih9A2D4A0CnsudjqQFg+ANAb2XOydICwPAHgMHKmpelBIDhDwDDK2NuTj0ADH8AGN205+dUA8DwB4DxTXOOTi0ADH8AmNy05ulUAsDwB4DiTGOuFh4Ahj8AFK/o+VpoABj+ADA9Rc7ZwgLA8AeA6Stq3lb+twECAOUrJACc/QNAeYqYuxMHgOEPAOWbdP5OFACGPwBUZ5I57B4AAAho7ABw9g8A1Rt3Ho8VAIY/ANTHOHPZJQAACGjkAHD2DwD1M+p8HikADH8AqK9R5rRLAAAQ0NAB4OwfAOpv2HltAwAAAQ0VAM7+AaA5hpnbNgAAENDAAHD2DwDNM2h+2wAAQEB9A8DZPwA0V785bgMAAAH1DABn/wDQfL3muQ0AAATUNQCc/QNAPrrNdRsAAAhIAABAQB0BYP0PAPlpn+82AAAQkAAAgIBaAsD6HwDytXLO2wAAQEACAAACEgAAENDxAHD9HwDytzzvbQBGtLh9Z9WHAAATEwBjEAEANJ0AGMHKwS8CAGiyVSm5/j+MxctP6XxMBADQQBvmti7ZAAzh+PDf9Y7OPxMBADSQABig48xfBACQAQHQw+Llp3Rd+6eURAAAjScAuug5+FcSAQA0mABoM9TwXyYCAGgoAdBm4/UPjfYJIgCABlrlVwA7iQAAcmcD0IMIACBnAqAPEQBArgTAIPMfH+3ju0QAANSNABjGBBGw8aqtBR8MAExOAAzLJgCAjAiAPjqu348SAaMGAwCUSACMapjB/vjHWP8DUFezVR9AI81/vPuKvwFn/Ys3XVv1IUCxbtwktmEMNgA9DPz1vfZhX+Phv3jTtcf/Azny67YwOgEwieWhX9Phb+gTiQiA0QiASfUY/lWvJA1+IhIBMDwB0EXTf4gY/kTW9O9fKIsAyIzhDyIAhiEApqCq9b/hD08QAdCfAMiE4Q+dRAD0JgDaNPEHhuEPvTXxexrKIAAKVvb63/CHwUQAdBIAQAgiAFoJgBWa9gPC2T+Mpmnf4zBNAgAIRQTAYwRAgap+9z9gOCIABMBxTfuBYP0Pk2na9zwUTQAAYYkAIhMABbH+h2YSAUQlAIDwRAARCYDkmx/wc4B4BEABrP8hDyKASAQAwAoigCjCB4BvdqCdnwtEED4AALoRAeROAEzI9X/IlwggZ6EDwDc3MIifE+QqdAAADEMEkCMBMAHrf4hDBJAbAQAwJBFATmarPoCqhPtGvnFT1UcAWZj0Z4fNIXVhAzAm38QANJkAAICAQgZAuPU/ALQJGQAAEJ0AGIPr/wA0XbgAsP4HgIABAAAIgJFZ/wOQAwEAAAGFCgDX/wHgMaECYFLW/wDkIuzfBcATDhyaTZf+8avTd394UsvjH3zzLemKS/f0/Lyrds6lT954dstjl53zQPrHX/vPNDPzxGMPHlibPvXVZ6XP7z4z7b5/fdp/cG1KaSmdetKjaf2TDqcz1x9M5535UDp/84/Si7bsS2dv+nGRTw+ALsIEgPV/b+tOOJKu+YWb05v+4iUtj1993UXp555zX9q8/mDH53x5z6b0tze1Dv8nrT2a/uytN7cM/3/7xuZ05adekB48sLbj37H3odVp70Mnprv2npK+eNcZKaWUfuWyb6U/edPXCnhWAPQT5xLAhddVfQS19tJn701vf8G3Wx57+JE16T2ffl7Hxx46sipdtXMuLS21Pv7e+dvSMzceOP7Pt373tLTtE5d2Hf4AVCvMBiCl9FgEfP01Y31qhOv/73/9/6Qbdp+Z7t//pOOPfe7rT0vX3fb09JqL7j3+2If//fz0rQee3PK5z3/WD9I7X/ytlsf+8HMXpsNHn2jMk9YeSe977a3pVRfcn04/+ZF04NBsunPv+nT9HWemv7/5men7K74uANMVZgOw8fJPPvY/bAJ6OuXER9OH33JLx+Pv+fTz0kOPrEkppXTH/evTRz5/bsufnzB7NH30rTenVTNPrAQOH12V/uubT2n5uHe/8hvpikv3pKed+pN0wuyxtGHd4XTJ2fvSe+e/nm79vevSn25dSGc8+ZEpPDOoD5cjqYswAdBCBPT0qgvuS29+3ndaHtv70Inp9z97cTq2NJN+8x/m0qNHW/9v8zuvvj2d85SHWx77wcMntJz9p5TS00/7Sc+vO7tqKW275O707ld+Y8JnAPUnAqiDmAGQ0kgREGH9v9IH3vi1dPrJh1oe++RNZ6crP/X8tHDPxpbHn/uMB9O7XnZnx79j7eyxjsf+8os/nfY9fGKxBwsNJQKoWqgAOH4ZYJlNQFcb1h1OH3xz66WApaWUdi6c1fLYmtXH0kffdnNavartbsCU0uknH+r47YFb7tmYLr56Pv3ixy9LH7nh3PSFu85I+w+uKf4JQEOIAKoUKgC6EgFdvfbie9PPX3xv34/5rVfckc7fvL/nn7/zxd/seOzQkdXpX29/anrfP12U3vixl6Rzfvf16WUfekW65vrz0g9/4rcFiEcEUBUBkFLfCIi2/l/pg2/6WjrtpMNd/+z8zfvTVZff0ffzf/3ld6atc/f0/ZhjSzPptu+dlv5g14XpZ98/n/7l9qeOfbzQVCKAKgiAZTYBHTY9+ZH0R2/4747HV69aSn/+tq+mNas7r/OvtGpmKX3s7V9Jf/eOL6VLt+xreYOgbvYfXJOu+JsXpTu/f8okhw2NJAIoW7gA6LgPYCUR0OF1z+28DPBTGw6ki5/xw6H/Ha++4L702Xf9R9p99WfSX//yl9OVL70zPf9ZP0izXe4dOHRkVfqrL50z0TFDU4kAyhQuAAYSAVNz+smH0msvvjdd/bpb0z//xg3pjqs/k37phXd3fNxt3zutgqODehABlEUAdPN4BES+/l+GDesOpw+95ZZ04pqjLY8/enTAtQLInAigDCEDoO9lgGU2ARN7w8dekm7YfWbH3xmw0sHDs+lI2xsGdfvLhyAaEcC0xfq7AEYwVCTQ11fu3pS+eNcZ6ayNB9IbfuY76UVb9qULNu9PG04+lA4eXp3+975T0wc+95x05FjrGf/Lnr23oiOGelncvtMmkqkRAF0Y/sW6Z3Fduub689I115838GPP2nig428lhKxcsm/ET7hyKocBIS8BpGTIl2HdCUdG+vjzNu9Pn/7VL3TcEwBA8WwA2giD4tz+vs+mr3779PSVu09Pt957Wvq/xXXp/v0npR8/MpuOLs2kdWuPpM2nHkzPeeqP0vyF30vzF93b9VcDASieAFjB8O90wuzRsW9GWrv6WLrsnAfSZec8UPBRATCpsJcA2hn+AEQSOgAMfQCiCh0Ay4QAANGEDwDDH4CIwgcAAEQkAAAgIAEAAAF5H4AoRn77UQByZgMAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACGi26gOgXvYs7NhV9TFAE22Z2zZf9THAKGwAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABeStgWng7U4AYbAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQLNVHwD1smdhx66qjwG62TK3bb7qY4Cc2AAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAALyVsC08HarADHYAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICAZqs+AOplz8KOXVUfAzTRlrlt81UfA4zCBgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJC3AqaFtzMFiMEGAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQ0W/UBUC97FnbsqvoYoJstc9vmqz4GyIkNAAAEJAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIG8FTAtvtwoQgw0AAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABCQAACGi26gOgXvYs7NhV9TFAE22Z2zZf9THAKGwAACAgAQAAAQkAAAhIAABAQAIAAAISAAAQkAAAgIAEAAAEJAAAICABAAABeStgWng7U4AYbAAAICABAAABCQAACEgAAEBAAgAAAhIAABCQAACAgAQAAAQkAAAgIAEAAAEJAAAISAAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQLNVHwD1smdhx66qjwG62TK3bb7qY4Cc2AAAQEACAAACEgAAEJAAAICABAAABCQAACAgAQAAAQkAAAhIAABAQAIAAALyVsC08HarADEIAICy3Lhp5E9Z3P7Yf2+8amvBB0N0AgBgGsYY9v0sbt/Z8ZgoYBICAKAoBQ/9QVZGgRhgVAIAYBIlD/1exACjEgAAo5pk6K8/d/TP2b97pA8XAwxDAAAMa9TBP86wH/bfM2QULMeAEKCdAAAYZNjBX9TAH+drDQgCIUA7bwQE0M8ww3/9ueUO/wmOodtvExCTDQBAN4MGf9UDv5eVx9VjK2AbQEo2AACd+g3/OpztD2vAsdoGxGYDALBs0OBvquVj77IRsA2IywYAIKV8h/9KtgGsIAAAeg3/Jq37h9XnOYmAWAQAEFu/4Z8zERCeewCAmCKs/AdZf677AgKzAQDiibTyH8QlgbAEABBL1JX/ICIgHAEAEH34L/M6hCIAgDi6nf0beq26vB62AHkSAEAMhv/wREAIfguAFnsWduyq+higaFse/e35jgcN//66/IbA4vadfjMgIzYAQNYM/wnYBGRNAABAQAIAyJaz/wLYAmRLAABZMvwLJAKyJACAGAz/yXj9siMAgOx0PfuncLYAzSYAgKxY/U+RSwFZEQBA3gz/Ynk9syEAgGxY/VfDFqCZBACQL2er0+F1zcLMhrmtS1UfBKNbvOnaqg8B6sV7/Zer7W2CU0reJrhhbACAPAUf/huvuGi6XyD465sDAQCQmeXhP/UIoNEEANB83db/QbUP/TIjwM2AzSIAgPxYT7eYWgR4nRtNAABkot+gdzmAdgIAaLb29X/Qs9JBA37xE7dN5wu3vd4uAzSHAABouMqGP40mAAAazPBnXAIAaK7gd//X9bq+ywDNIACAfAS9/t9LaWf/XvdGEgAADWT1z6QEAEDDGP4UQQAANIjhT1EEANBMAW8AbNLwdyNg/QkAIA+Z34hW1zv+j8v89c+RAADIQJ3O/mkGAQBQc01a/dMcAgCgxgx/pkUAANSU4c80CQCAGjL8mTYBADCGad6VX/s7/smCAAAY0fKArmpQO/unCAIAYATtQ7/oCLD6pywCAGBIvYZzURFg+FMmAQAwhGmv+w1/yiYAAIYwaABPEgiGP1UQAABDmkYEuOOfqggAIA/7d5fyZaa5CRjn69VGSa8/xREAQDNdsq+yL11UBOS8+t941daqD4EBBADAFAwa7jkPf5pBAACMYZgBPe6vDRr+lEEAAIxpnAgw/KkLAQDko4Ib0UYZ2Nne8e8GwEYSAEBzVXgj4EpF3RSYy9m/GwCbQQAAFGDSCMhl+NMcAgCgIOMOccOfKggAoNnaLwNUfD161GHe+OHf9npb/zeHAGiojS+8supDACbU+OFPowkAID8N2wI0lrv/G00AAM1Xk98GWKk9Agb9cw6s/5tFADSYywBQb8tDvtd/Q5UEAJCnmqynsz3zr8nry/gEQMPZAsDjangZIBLr/+YRABkQAdCDs9Tp8LpmQQBkQgRAsgWoiLP/ZhIAGREB0IWz1WJ5PbMhADIjAgiv2xbA0CpGl9fR2X9zCYAMiQDCcymgFIZ/swmATIkAaGMLMBmvX3ZmNsxtXar6IJiuxZuurfoQoBo3bup8bP255R9H01n9Z0kABCMGCEcETMbwz9Zs1QdAuVwaIJrF7VUfAdSTewCArHU9W3U9ezjO/rMmAIDsiYAxGP7ZEwBACCJgBIZ/CAIACEMEDMHwD0MAAIiAx3gdQhEAQCg9z2ajD78ez9/Zf74EABBO3wiIFgJ9nrPhnzdvBASEtrh9Z/c/iPBmQQZ/aDYAQGhhLwkY/uEJACC8UJcErPx5nAAASAOGXy4R0Od5GP7xuAcAoE3P+wJSaua9AQY/XdgAALQZuA1oykZgwLEa/rHZAAD00XcbsKxOW4Eh4sTgJyUbAIC+hhqWddgKDHkMhj/LbAAAhjTUNmClaW4GRgwOg592AgBgRCOHwErjRMEE2wWDn14EAMAEJoqBKTH0GYYAAChIlTFg6DOqmZRSEgEAxSojBgx9JiEAAEoySRQY9hRNAABAQN4HAAACEgAAEJAAAICABAAABLQqpZQeXNg5U/WBAADleHBh54wNAAAEJAAAICABAAABHQ8A9wEAQP6W570NAAAEJAAAICABAAABtQSA+wAAIF8r57wNAAAEJAAAIKCOAHAZAADy0z7fbQAAICABAAABdQ0AlwEAIB/d5roNAAAE1DMAbAEAoPl6zXMbAAAIqG8A2AIAQHP1m+M2AAAQ0MAAsAUAgOYZNL9tAAAgoKECwBYAAJpjmLltAwAAAQ0dALYAAFB/w85rGwAACGikALAFAID6GmVOj7wBEAEAUD+jzmeXAAAgoLECwBYAAOpjnLk89gZABABA9cadxy4BAEBAEwWALQAAVGeSOTzxBkAEAED5Jp2/hVwCEAEAUJ4i5q57AAAgoMICwBYAAKavqHlb6AZABADA9BQ5Zwu/BCACAKB4Rc/XqdwDIAIAoDjTmKtTuwlQBADA5KY1T6f6WwAiAADGN805OvVfAxQBADC6ac/PUt4HQAQAwPDKmJulvRGQCACAwcqal6W+E6AIAIDeypyTpb8VsAgAgE5lz8dKh/GGua1LVX59AKhaVSfGlf5lQLYBAERW5Rys/G8DFAEARFT1/KvV8HVJAIDcVT34l1W+AVipLi8KAExDneZcbQ6knW0AALmo0+BfVqsNwEp1fLEAYFR1nWe1PKh2tgEANE1dB/+yWh9cOyEAQN3VffAva8RBthMCANRNUwb/skYdbDshAEDVmjb4lzXyoNsJAQDK1tTBv6zRB99OCAAwbU0f/MuyeBLdiAEAipLL0F8puyfUjRgAYFQ5Dv2Vsn5yvQgCANrlPvDbhXqyvQgCgHiiDfx2/w8eWPS/OI7SkAAAAABJRU5ErkJggg=="

col_logo, col_title = st.columns([1, 8])
with col_logo:
    st.markdown(
        f'''<img src="data:image/png;base64,{LOGO_B64}"
             width="80" style="border-radius:14px; margin-top:4px;">''',
        unsafe_allow_html=True
    )
with col_title:
    st.markdown(
        """<h1 style="margin:0; padding-top:8px; color:#12345A;
                      font-size:2.2rem; font-weight:800; letter-spacing:-0.5px;">
            Yellow Sheet
           </h1>
           <p style="margin:0; color:#5a7a9a; font-size:0.95rem; font-weight:500;">
            Student Exam Data Consolidator &nbsp;·&nbsp; Result Generator &nbsp;·&nbsp; PDF Slips
           </p>""",
        unsafe_allow_html=True
    )
st.markdown("---")

# ── Faculty / Subject Definitions ─────────────────────────────────────────────
# Each faculty defines:
#   'subjects'  : ordered list of (abbr, full_name, annual_max, internal_max)
#   'optional'  : list of abbrs that are optional (student picks exactly one)
#                 the app auto-detects which optional the student took
#   'core'      : always-present subjects (first 5 or first 6 excl optional group)

FACULTY_CONFIG = {
    "Arts": {
        "core":     ["ENG", "MAR", "GEO", "PSY", "ECO"],
        "optional": ["SOC", "VOC"],          # student takes exactly ONE
        "subjects": {
            "ENG": ("English",     80, 20),
            "MAR": ("Marathi",     80, 20),
            "GEO": ("Geography",   80, 20),
            "PSY": ("Psychology",  80, 20),
            "ECO": ("Economics",   80, 20),
            "SOC": ("Sociology",   80, 20),
            "VOC": ("Vocational",  80, 20),
        },
    },
    "Commerce": {
        "core":     ["ENG", "MAR", "ECO", "ACC", "O.C.", "S.P."],
        "optional": [],
        "subjects": {
            "ENG":  ("English",    80, 20),
            "MAR":  ("Marathi",    80, 20),
            "ECO":  ("Economics",  80, 20),
            "ACC":  ("Accounts",   80, 20),
            "O.C.": ("O.C.",       80, 20),
            "S.P.": ("S.P.",       80, 20),
        },
    },
    "Science": {
        "core":     ["ENG", "MAR", "GEO", "PHY", "CHE"],
        "optional": ["BIO", "MATH"],         # student takes exactly ONE
        "subjects": {
            "ENG":  ("English",    80, 20),
            "MAR":  ("Marathi",    80, 20),
            "GEO":  ("Geography",  80, 20),
            "PHY":  ("Physics",    80, 20),
            "CHE":  ("Chemistry",  80, 20),
            "BIO":  ("Biology",    70, 30),
            "MATH": ("Maths",      70, 30),
        },
    },
}

# Passing marks reference (individual exams — informational only)
# Final overall pass = average/100 >= 35 in ALL 6 subjects
EXAM_PASS = {
    "FIRST UNIT TEST (25)":  9,
    "FIRST TERM EXAM (50)": 18,
    "SECOND UNIT TEST (25)": 9,
    "ANNUAL EXAM (70/80)":  28,   # 28 for /80, 25 for /70 — shown only
}

def custom_round(x):
    try:
        return int(np.floor(float(x) + 0.5))
    except:
        return 0

def clean_marks(val):
    if isinstance(val, str):
        v = val.strip().upper()
        if v in ("AB", ""):
            return 0.0
    try:
        return float(val)
    except:
        return 0.0

# ── Detect which 6 subjects a student actually has (handles optional) ─────────
def detect_student_subjects(faculty, df_row, cfg):
    """
    Returns ordered list of 6 abbrs for this student.
    For optional groups, picks whichever column has a non-zero / non-blank value.
    If both blank, defaults to first optional.
    """
    core = cfg["core"]
    opt  = cfg["optional"]
    if not opt:
        return core  # exactly 6 core subjects

    # Pick optional: whichever has a value in this row
    chosen_opt = opt[0]  # default
    for o in opt:
        val = str(df_row.get(o, "")).strip()
        if val and val.upper() not in ("", "NAN", "0"):
            chosen_opt = o
            break
    return core + [chosen_opt]

# ── PDF Generation ────────────────────────────────────────────────────────────
def build_exam_pdf(school_name, faculty_name, exam_label, student_results,
                   cfg, pos_cols, selected_exam_data):
    """
    Landscape A4.  2 slips per page drawn directly with canvas so each slip
    fills the FULL page height — no empty space at the bottom.
    Signature line is always pinned to the very bottom of the slip.
    """
    from reportlab.lib.pagesizes import landscape
    from reportlab.pdfgen         import canvas as rl_canvas
    from reportlab.platypus       import Frame, KeepInFrame

    exam_meta = {
        "FIRST UNIT TEST (25)":  {"max_per_sub": 25,  "pass_mark": 9,  "total_max": 150},
        "FIRST TERM EXAM (50)":  {"max_per_sub": 50,  "pass_mark": 18, "total_max": 300},
        "SECOND UNIT TEST (25)": {"max_per_sub": 25,  "pass_mark": 9,  "total_max": 150},
        "ANNUAL EXAM (70/80)":   {"max_per_sub": None,"pass_mark": 28, "total_max": None},
    }
    meta = exam_meta.get(exam_label,
                         {"max_per_sub": None, "pass_mark": None, "total_max": None})

    # Clean display title printed on every slip header
    exam_display_title = {
        "FIRST UNIT TEST (25)":  "FIRST UNIT TEST RESULT",
        "FIRST TERM EXAM (50)":  "FIRST TERM EXAM RESULT",
        "SECOND UNIT TEST (25)": "SECOND UNIT TEST RESULT",
        "ANNUAL EXAM (70/80)":   "ANNUAL EXAM RESULT",
    }.get(exam_label, exam_label)

    PAGE     = landscape(A4)          # (841.9, 595.3) points  ≈ 297×210 mm
    PW, PH   = PAGE
    MARGIN   = 10 * mm
    GAP      = 6  * mm               # divider gap between the two slips
    SLIP_W   = (PW - 2*MARGIN - GAP) / 2
    SLIP_H   = PH - 2*MARGIN         # FULL usable height

    # Row heights for the marks table — tall rows for readability
    ROW_H    = 9 * mm

    # ── Styles ────────────────────────────────────────────────────────────────
    school_style = ParagraphStyle("sch", fontSize=13, fontName="Helvetica-Bold",
                                  alignment=TA_CENTER, spaceAfter=3,
                                  leading=17, wordWrap="LTR",
                                  textColor=colors.Color(0.12, 0.31, 0.49))
    exam_style   = ParagraphStyle("exm", fontSize=13, fontName="Helvetica-Bold",
                                  alignment=TA_CENTER, spaceAfter=1,
                                  textColor=colors.Color(0.75, 0.15, 0.0))
    fac_style    = ParagraphStyle("fac", fontSize=11, fontName="Helvetica",
                                  alignment=TA_CENTER, spaceAfter=4)
    lbl_style    = ParagraphStyle("lbl", fontSize=11, fontName="Helvetica-Bold")
    val_style    = ParagraphStyle("val", fontSize=11, fontName="Helvetica")
    res_pass     = ParagraphStyle("rp",  fontSize=16, fontName="Helvetica-Bold",
                                  textColor=colors.Color(0.05,0.55,0.05),
                                  alignment=TA_CENTER)
    res_fail     = ParagraphStyle("rf",  fontSize=16, fontName="Helvetica-Bold",
                                  textColor=colors.Color(0.8,0.1,0.1),
                                  alignment=TA_CENTER)
    sig_style    = ParagraphStyle("sg",  fontSize=10, fontName="Helvetica",
                                  alignment=TA_CENTER)

    S_COL = SLIP_W * 0.52
    M_COL = SLIP_W * 0.24
    O_COL = SLIP_W * 0.24

    def slip_content(sr):
        """Return (top_elems, result_text, is_pass) — signature drawn separately."""
        roll   = sr["roll"]
        name   = sr["name"]
        subj_6 = sr["subj_6"]
        exam_d = selected_exam_data.get(roll, {})
        is_ann = (exam_label == "ANNUAL EXAM (70/80)")
        elems  = []

        # ── Header block ──────────────────────────────────────────────────────
        elems.append(Spacer(1, 3*mm))
        elems.append(Paragraph(school_name, school_style))
        elems.append(Paragraph(exam_display_title, exam_style))
        elems.append(Paragraph(f"{faculty_name} Faculty", fac_style))
        elems.append(HRFlowable(width="100%", thickness=2,
                                color=colors.Color(0.12, 0.31, 0.49),
                                spaceAfter=4*mm))

        # ── Roll / Name ───────────────────────────────────────────────────────
        info = Table(
            [[Paragraph("Roll No. :", lbl_style),
              Paragraph(str(roll), val_style),
              Paragraph("Name :", lbl_style),
              Paragraph(str(name), val_style)]],
            colWidths=[SLIP_W*0.17, SLIP_W*0.17, SLIP_W*0.13, SLIP_W*0.53],
        )
        info.setStyle(TableStyle([
            ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ("TOPPADDING",    (0,0),(-1,-1), 5),
        ]))
        elems.append(info)
        elems.append(Spacer(1, 4*mm))

        # ── Marks table ───────────────────────────────────────────────────────
        tbl_data  = [["Subject", "Max Marks", "Obtained"]]
        total_obt = 0
        total_max = 0

        for abbr in subj_6:
            subj_name, ann_max, _ = cfg["subjects"][abbr]
            max_m = ann_max if is_ann else meta["max_per_sub"]
            raw   = exam_d.get(abbr, "")
            try:
                obt = int(float(raw)) if str(raw).strip().upper() != "AB" else "AB"
            except:
                obt = str(raw)
            if isinstance(obt, int):
                total_obt += obt
                total_max += max_m
            tbl_data.append([subj_name, str(max_m), str(obt)])

        tbl_data.append(["TOTAL", str(total_max),
                          str(total_obt) if isinstance(total_obt, int) else "-"])
        pct = round(total_obt / total_max * 100, 2) \
              if isinstance(total_obt, int) and total_max else "-"
        tbl_data.append(["Percentage", "",
                          f"{pct} %" if pct != "-" else "-"])

        n_rows   = len(tbl_data)
        row_heights = [ROW_H] * n_rows          # uniform tall rows
        marks_tbl = Table(tbl_data,
                          colWidths=[S_COL, M_COL, O_COL],
                          rowHeights=row_heights)
        marks_tbl.setStyle(TableStyle([
            # Header
            ("BACKGROUND",    (0,0),(-1,0),    colors.Color(0.12,0.31,0.49)),
            ("TEXTCOLOR",     (0,0),(-1,0),    colors.white),
            ("FONTNAME",      (0,0),(-1,0),    "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1),   11),
            ("ALIGN",         (1,0),(-1,-1),   "CENTER"),
            ("ALIGN",         (0,0),(0,-1),    "LEFT"),
            ("VALIGN",        (0,0),(-1,-1),   "MIDDLE"),
            # Alternating subject rows
            ("ROWBACKGROUNDS",(0,1),(-1,n_rows-3),
                              [colors.white, colors.Color(0.94,0.96,0.99)]),
            # Total row
            ("BACKGROUND",    (0,-2),(-1,-2),  colors.Color(0.80,0.80,0.80)),
            ("FONTNAME",      (0,-2),(-1,-2),  "Helvetica-Bold"),
            # Percentage row
            ("BACKGROUND",    (0,-1),(-1,-1),  colors.Color(0.88,0.92,0.97)),
            ("SPAN",          (0,-1),(1,-1)),
            ("FONTNAME",      (0,-1),(-1,-1),  "Helvetica-Bold"),
            # Grid
            ("GRID",          (0,0),(-1,-1),   0.6, colors.grey),
            ("LEFTPADDING",   (0,0),(-1,-1),   5),
            ("RIGHTPADDING",  (0,0),(-1,-1),   5),
        ]))
        elems.append(marks_tbl)
        elems.append(Spacer(1, 5*mm))

        # ── Pass / Fail ───────────────────────────────────────────────────────
        indiv_pass = True
        for abbr in subj_6:
            raw = exam_d.get(abbr, "")
            try:
                m   = float(raw)
                req = (28 if cfg["subjects"][abbr][1] == 80 else 25) \
                      if is_ann else meta["pass_mark"]
                if m < req:
                    indiv_pass = False; break
            except:
                indiv_pass = False

        # Result in a shaded box
        res_text  = "✓   PASS" if indiv_pass else "✗   FAIL"
        res_fill  = colors.Color(0.85, 0.97, 0.85) if indiv_pass \
                    else colors.Color(0.99, 0.87, 0.87)
        res_border= colors.Color(0.05,0.55,0.05) if indiv_pass \
                    else colors.Color(0.8,0.1,0.1)
        res_tbl   = Table([[Paragraph(res_text,
                                       res_pass if indiv_pass else res_fail)]],
                          colWidths=[SLIP_W * 0.96])
        res_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,-1), res_fill),
            ("BOX",           (0,0),(-1,-1), 1.5, res_border),
            ("BOTTOMPADDING", (0,0),(-1,-1), 8),
            ("TOPPADDING",    (0,0),(-1,-1), 8),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ]))
        elems.append(res_tbl)

        return elems, indiv_pass

    # ── Canvas-based renderer so slips fill full page height ─────────────────
    buf = BytesIO()
    c   = rl_canvas.Canvas(buf, pagesize=PAGE)

    def draw_slip_on_canvas(c, sr, x_offset):
        """
        Draw one slip inside a Frame at x_offset.
        Top section fills normally; signature is pinned to the bottom.
        """
        SIG_H   = 18 * mm      # reserved height at bottom for signatures
        INNER_H = SLIP_H - SIG_H

        # Outer border box
        c.setStrokeColor(colors.Color(0.12, 0.31, 0.49))
        c.setLineWidth(1.5)
        c.rect(x_offset, MARGIN, SLIP_W, SLIP_H, stroke=1, fill=0)

        # ── Top content frame ─────────────────────────────────────────────────
        top_elems, indiv_pass = slip_content(sr)

        top_frame = Frame(
            x_offset + 2*mm,
            MARGIN + SIG_H,
            SLIP_W - 4*mm,
            INNER_H,
            leftPadding=0, rightPadding=0,
            topPadding=0,  bottomPadding=0,
            showBoundary=0,
        )
        kif = KeepInFrame(SLIP_W - 4*mm, INNER_H, top_elems,
                          mode='shrink')
        top_frame.addFromList([kif], c)

        # ── Divider line above signature ──────────────────────────────────────
        sig_y = MARGIN + SIG_H
        c.setStrokeColor(colors.Color(0.5, 0.5, 0.5))
        c.setLineWidth(0.5)
        c.line(x_offset + 3*mm, sig_y, x_offset + SLIP_W - 3*mm, sig_y)

        # ── Signature labels pinned to bottom ─────────────────────────────────
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        quarter = SLIP_W / 4
        # Left signature
        sig_line_y = MARGIN + 12*mm
        c.line(x_offset + quarter*0.3, sig_line_y,
               x_offset + quarter*1.7, sig_line_y)
        c.drawCentredString(x_offset + quarter,
                            MARGIN + 4*mm, "Class Teacher")
        # Right signature
        c.line(x_offset + quarter*2.3, sig_line_y,
               x_offset + quarter*3.7, sig_line_y)
        c.drawCentredString(x_offset + quarter*3,
                            MARGIN + 4*mm, "Principal")

    for i in range(0, len(student_results), 2):
        # Left slip
        draw_slip_on_canvas(c, student_results[i], MARGIN)

        # Right slip (or blank box if odd number of students)
        right_x = MARGIN + SLIP_W + GAP
        if i + 1 < len(student_results):
            draw_slip_on_canvas(c, student_results[i+1], right_x)
        else:
            # Empty right slot — just draw border
            c.setStrokeColor(colors.Color(0.7, 0.7, 0.7))
            c.setLineWidth(1)
            c.rect(right_x, MARGIN, SLIP_W, SLIP_H, stroke=1, fill=0)

        # Vertical cut line between slips
        mid_x = MARGIN + SLIP_W + GAP / 2
        c.setDash(4, 3)
        c.setStrokeColor(colors.Color(0.6, 0.6, 0.6))
        c.setLineWidth(0.5)
        c.line(mid_x, MARGIN, mid_x, MARGIN + SLIP_H)
        c.setDash()

        if i + 2 < len(student_results):
            c.showPage()

    c.save()
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
uploaded_file = st.file_uploader("Upload Excel Marksheet", type="xlsx")

if uploaded_file:
    try:
        xl = pd.ExcelFile(uploaded_file)

        # ── Faculty selection ──────────────────────────────────────────────────
        st.markdown("---")
        faculty = st.selectbox("🎓 Select Faculty", list(FACULTY_CONFIG.keys()))
        cfg     = FACULTY_CONFIG[faculty]

        # Show subject table
        st.markdown("**Subject Configuration:**")
        sub_info = []
        for abbr in cfg["core"] + cfg["optional"]:
            name, am, im = cfg["subjects"][abbr]
            tag = ""
            if abbr in cfg["optional"]:
                tag = " *(optional)*"
            sub_info.append({
                "Abbr": abbr, "Subject": name + tag,
                "Annual Max": am, "Internal Max": im,
                "Total": f"25+50+25+{am}+{im}=200"
            })
        st.dataframe(pd.DataFrame(sub_info), hide_index=True, use_container_width=True)

        # ── Parse exam sheets ──────────────────────────────────────────────────
        exam_configs = [
            {"label": "FIRST UNIT TEST (25)",  "sheets": ["FIRST UNIT TEST"]},
            {"label": "FIRST TERM EXAM (50)",  "sheets": ["FIRST TERM"]},
            {"label": "SECOND UNIT TEST (25)", "sheets": ["SECOND UNIT TEST"]},
            {"label": "ANNUAL EXAM (70/80)",   "sheets": ["ANNUAL EXAM"]},
        ]

        all_students = {}   # roll -> {Name, Exams, subjects (list of 6 abbrs)}

        for config in exam_configs:
            sheet_name = next(
                (s for s in xl.sheet_names if s.strip().upper() in config["sheets"]), None
            )
            if not sheet_name:
                continue

            df = xl.parse(sheet_name)
            df.columns = df.columns.astype(str).str.strip().str.upper()

            # Normalise O.C. and S.P. column names (dots may be stripped)
            col_map = {}
            for c in df.columns:
                cn = c.replace(" ", "")
                if cn in ("OC", "O.C"):  col_map[c] = "O.C."
                if cn in ("SP", "S.P"):  col_map[c] = "S.P."
            if col_map:
                df = df.rename(columns=col_map)

            t_col = next((c for c in df.columns if "TOTAL" in c), None)
            p_col = next((c for c in df.columns if "%" in c or "PERCENT" in c), None)
            r_col = next((c for c in df.columns if "RESULT" in c), None)

            for _, row in df.iterrows():
                roll = str(row.get("ROLL NO.", "")).strip()
                if not roll or roll.lower() == "nan":
                    continue

                if roll not in all_students:
                    subj_6 = detect_student_subjects(faculty, row, cfg)
                    all_students[roll] = {
                        "Name":     str(row.get("STUDENT NAME", "Unknown")),
                        "Exams":    {},
                        "subjects": subj_6,
                    }

                subj_6 = all_students[roll]["subjects"]
                marks  = {}
                for abbr in subj_6:
                    raw = str(row.get(abbr, "0")).strip()
                    marks[abbr] = raw if raw.upper() == "AB" else row.get(abbr, 0)

                try:
                    marks["Grand Total"] = str(row.get(t_col, "")) if t_col else ""
                except:
                    marks["Grand Total"] = ""
                try:
                    raw_p = row.get(p_col, "")
                    marks["%"] = str(round(float(raw_p), 2)) if str(raw_p).strip() else ""
                except:
                    marks["%"] = ""
                marks["Result"] = str(row.get(r_col, ""))
                all_students[roll]["Exams"][config["label"]] = marks

        if not all_students:
            st.error("No student data found. Check sheet names match: FIRST UNIT TEST, FIRST TERM, SECOND UNIT TEST, ANNUAL EXAM")
            st.stop()

        student_rolls = sorted(
            all_students.keys(),
            key=lambda x: float(x) if x.replace(".", "", 1).isdigit() else 0
        )

        categories = [
            "FIRST UNIT TEST (25)",
            "FIRST TERM EXAM (50)",
            "SECOND UNIT TEST (25)",
            "ANNUAL EXAM (70/80)",
            "INT/PRACTICAL (20/30)",
            "Total Marks Out of 200",
            "Average Marks 200/2=100",
        ]
        result_cols = ["Grand Total", "%", "Result", "Remark", "Rank"]

        # ── Internal Marks Input ───────────────────────────────────────────────
        st.markdown("---")
        st.subheader("📝 Enter Internal / Practical Marks")
        st.info("Enter marks for each student. Subject columns match each student's chosen optional subject.")

        if "internal_marks" not in st.session_state:
            st.session_state.internal_marks = {
                roll: {abbr: "0" for abbr in all_students[roll]["subjects"]}
                for roll in student_rolls
            }

        # Header row
        hdr = st.columns([0.6, 1.8] + [0.9]*6)
        hdr[0].markdown("**Roll**")
        hdr[1].markdown("**Name**")

        for roll in student_rolls:
            subj_6 = all_students[roll]["subjects"]
            name   = all_students[roll]["Name"]
            cols   = st.columns([0.6, 1.8] + [0.9]*6)
            cols[0].write(roll)
            cols[1].write(name)
            for i, abbr in enumerate(subj_6):
                _, am, im = cfg["subjects"][abbr]
                val = cols[i+2].text_input(
                    label=f"{roll}-{abbr}",
                    value=st.session_state.internal_marks[roll].get(abbr, "0"),
                    key=f"int_{roll}_{abbr}",
                    label_visibility="collapsed",
                    placeholder=f"{abbr} /{im}",
                )
                st.session_state.internal_marks[roll][abbr] = val

        # ── Build base_df (one universal subject slot per position) ────────────
        # Since different students may have different optional subjects,
        # we use positional column names Sub1-Sub6 for the dataframe
        # but store actual abbr names in subject headers for display

        # Determine display subject headers — use most common subject set
        # (since all students in one faculty have same core, only optional differs)
        display_subj = cfg["core"] + [cfg["optional"][0]] if cfg["optional"] else cfg["core"]

        pos_cols  = [f"Sub{i+1}" for i in range(6)]   # positional df columns
        base_rows = []

        for roll in student_rolls:
            s      = all_students[roll]
            subj_6 = s["subjects"]

            for cat in categories:
                row_data = {
                    "Roll No.": roll if cat == "FIRST UNIT TEST (25)" else "",
                    "Column1":  s["Name"] if cat == "FIRST UNIT TEST (25)" else "",
                    "Column2":  cat,
                    "_subjects": "|".join(subj_6),   # hidden metadata
                }
                for pos, pc in enumerate(pos_cols):
                    row_data[pc] = ""
                for rc in result_cols:
                    row_data[rc] = ""

                if cat in s["Exams"]:
                    exam_marks = s["Exams"][cat]
                    for pos, abbr in enumerate(subj_6):
                        row_data[pos_cols[pos]] = str(exam_marks.get(abbr, ""))
                    row_data["Grand Total"] = exam_marks.get("Grand Total", "")
                    row_data["%"]           = exam_marks.get("%", "")
                    row_data["Result"]      = exam_marks.get("Result", "")

                elif cat == "INT/PRACTICAL (20/30)":
                    for pos, abbr in enumerate(subj_6):
                        row_data[pos_cols[pos]] = st.session_state.internal_marks[roll].get(abbr, "0")

                base_rows.append(row_data)

        base_df = pd.DataFrame(base_rows)
        for col in base_df.columns:
            if col != "_subjects":
                base_df[col] = base_df[col].astype(str).replace("nan", "")

        # Inject latest internal marks
        for i, roll in enumerate(student_rolls):
            subj_6 = all_students[roll]["subjects"]
            for pos, abbr in enumerate(subj_6):
                base_df.at[i*7 + 4, pos_cols[pos]] = \
                    st.session_state.internal_marks[roll].get(abbr, "0")

        # Display with actual subject names as column headers
        st.markdown("---")
        st.subheader("📊 Marks Preview & Edit")

        display_df = base_df.drop(columns=["_subjects"]).copy()
        # Rename Sub1-Sub6 to actual subject names for display
        rename_map = {pos_cols[i]: display_subj[i] if i < len(display_subj) else pos_cols[i]
                      for i in range(6)}
        display_df = display_df.rename(columns=rename_map)
        edited_display = st.data_editor(display_df, hide_index=True, use_container_width=True)

        # Map edited values back to positional columns
        rev_rename = {v: k for k, v in rename_map.items()}
        edited_df  = edited_display.rename(columns=rev_rename)

        # ── Generate Report ────────────────────────────────────────────────────
        if st.button("🚀 Generate Final Report & Rank"):

            student_results = []

            for s_idx, roll in enumerate(student_rolls):
                subj_6 = all_students[roll]["subjects"]
                block  = edited_df.iloc[s_idx*7 : s_idx*7+7].copy().reset_index(drop=True)

                raw = {}
                for row_i in range(5):
                    raw[row_i] = {pc: clean_marks(block.at[row_i, pc]) for pc in pos_cols}

                t200 = {pc: sum(raw[r][pc] for r in range(5)) for pc in pos_cols}
                a100 = {pc: custom_round(t200[pc] / 2) for pc in pos_cols}
                gt   = sum(a100.values())
                pc_  = round((gt / 600) * 100, 2)
                isp  = all(a100[pc] >= 35 for pc in pos_cols)

                student_results.append({
                    "roll":   roll,
                    "name":   all_students[roll]["Name"],
                    "subj_6": subj_6,
                    "t200":   t200,
                    "a100":   a100,
                    "gt":     gt,
                    "pc":     pc_,
                    "pass":   isp,
                    "rank":   "",
                })

            # Dense rank — PASS students only
            pass_gts = sorted(set(sr["gt"] for sr in student_results if sr["pass"]), reverse=True)
            rank_map = {gt_val: r+1 for r, gt_val in enumerate(pass_gts)}
            for sr in student_results:
                sr["rank"] = rank_map[sr["gt"]] if sr["pass"] else ""

            # Rebuild final_df
            processed = []
            for s_idx, sr in enumerate(student_results):
                block = edited_df.iloc[s_idx*7 : s_idx*7+7].copy().reset_index(drop=True)
                for pc in pos_cols:
                    block.at[5, pc] = str(int(sr["t200"][pc]))
                    block.at[6, pc] = str(int(sr["a100"][pc]))
                block.at[6, "Grand Total"] = str(sr["gt"])
                block.at[6, "%"]           = str(sr["pc"])
                block.at[6, "Result"]      = "PASS" if sr["pass"] else "FAIL"
                block.at[6, "Rank"]        = str(sr["rank"])
                processed.append(block)

            final_df = pd.concat(processed).reset_index(drop=True)

            # ── Save everything to session_state so PDF section persists ──────
            st.session_state.report_ready       = True
            st.session_state.student_results    = student_results
            st.session_state.final_df           = final_df
            st.session_state.all_students_snap  = all_students   # snapshot for PDF
            st.session_state.faculty_snap       = faculty
            st.session_state.cfg_snap           = cfg
            st.session_state.pos_cols_snap      = pos_cols
            st.session_state.exam_configs_snap  = exam_configs
            st.session_state.xl_sheet_names     = xl.sheet_names
            st.session_state.display_subj_snap  = display_subj

        # ── Show report if already generated ──────────────────────────────────
        if st.session_state.get("report_ready"):
            student_results = st.session_state.student_results
            final_df        = st.session_state.final_df
            cfg_s           = st.session_state.cfg_snap
            display_subj_s  = st.session_state.display_subj_snap

            passed = sum(1 for sr in student_results if sr["pass"])
            st.success(f"✅ {passed} PASS  |  {len(student_results)-passed} FAIL")

            summary_rows = []
            for sr in student_results:
                row_s = {"Roll No.": sr["roll"], "Name": sr["name"]}
                for i, abbr in enumerate(sr["subj_6"]):
                    row_s[f"{abbr} /100"] = sr["a100"][pos_cols[i]]
                row_s["Grand Total"] = sr["gt"]
                row_s["%"]           = sr["pc"]
                row_s["Result"]      = "PASS" if sr["pass"] else "FAIL"
                row_s["Rank"]        = sr["rank"]
                summary_rows.append(row_s)

            st.subheader("📋 Result Summary")
            st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

            # ── Build Excel ────────────────────────────────────────────────────
            wb   = Workbook()
            ws   = wb.active
            ws.title = "Consolidated"

            ws_h = wb.create_sheet("_RankHelper")
            ws_h.sheet_state = "hidden"
            ws_h.cell(row=1, column=1, value="GT")
            ws_h.cell(row=1, column=2, value="IsPass")

            hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            hdr_fill = PatternFill("solid", start_color="1F4E79")
            cat_fill = {
                "FIRST UNIT TEST (25)":    PatternFill("solid", start_color="DDEBF7"),
                "FIRST TERM EXAM (50)":    PatternFill("solid", start_color="E2EFDA"),
                "SECOND UNIT TEST (25)":   PatternFill("solid", start_color="FFF2CC"),
                "ANNUAL EXAM (70/80)":     PatternFill("solid", start_color="FCE4D6"),
                "INT/PRACTICAL (20/30)":   PatternFill("solid", start_color="EAD1DC"),
                "Total Marks Out of 200":  PatternFill("solid", start_color="D9D9D9"),
                "Average Marks 200/2=100": PatternFill("solid", start_color="BDD7EE"),
            }
            thin = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"),  bottom=Side(style="thin"),
            )
            ctr = Alignment(horizontal="center", vertical="center")

            col_headers = (
                ["Roll No.", "Student Name", "Exam Type"]
                + [cfg_s["subjects"][a][0] if a in cfg_s["subjects"] else a
                   for a in display_subj_s]
                + result_cols
            )
            for ci, h in enumerate(col_headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font=hdr_font; c.fill=hdr_fill; c.alignment=ctr; c.border=thin

            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 28
            for i in range(6):
                ws.column_dimensions[get_column_letter(4+i)].width = 11
            for i in range(len(result_cols)):
                ws.column_dimensions[get_column_letter(10+i)].width = 13

            SUB_S = 4
            GT_C  = SUB_S + 6
            PCT_C = GT_C  + 1
            RES_C = PCT_C + 1
            REM_C = RES_C + 1
            RNK_C = REM_C + 1

            sub_lets = [get_column_letter(SUB_S + i) for i in range(6)]
            gt_let   = get_column_letter(GT_C)
            res_let  = get_column_letter(RES_C)
            n        = len(student_results)
            h_gt_rng = f"_RankHelper!$A$2:$A${n+1}"
            avg_excel_rows = []

            for s_idx, roll in enumerate([sr["roll"] for sr in student_results]):
                sr   = student_results[s_idx]
                brow = 2 + s_idx * 7

                for cat_idx, cat in enumerate(categories):
                    erow = brow + cat_idx
                    fl   = cat_fill.get(cat, PatternFill("solid", start_color="FFFFFF"))

                    ws.cell(row=erow, column=1, value=roll if cat_idx == 0 else "")
                    ws.cell(row=erow, column=2, value=sr["name"] if cat_idx == 0 else "")
                    ws.cell(row=erow, column=3, value=cat)

                    if cat == "Total Marks Out of 200":
                        r1, r5 = brow, brow + 4
                        for i, sl in enumerate(sub_lets):
                            c = ws.cell(row=erow, column=SUB_S+i,
                                        value=f"=SUM({sl}{r1}:{sl}{r5})")
                            c.fill=fl; c.border=thin; c.alignment=ctr
                            c.font=Font(name="Arial", bold=True)
                        for ri in range(len(result_cols)):
                            c = ws.cell(row=erow, column=GT_C+ri, value="")
                            c.fill=fl; c.border=thin

                    elif cat == "Average Marks 200/2=100":
                        trow = erow - 1
                        for i, sl in enumerate(sub_lets):
                            c = ws.cell(row=erow, column=SUB_S+i,
                                        value=f"=ROUND({sl}{trow}/2,0)")
                            c.fill=fl; c.border=thin; c.alignment=ctr
                            c.font=Font(name="Arial", bold=True)
                        c = ws.cell(row=erow, column=GT_C,
                                    value=f"=SUM({sub_lets[0]}{erow}:{sub_lets[-1]}{erow})")
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        c.font=Font(name="Arial", bold=True, color="1F4E79")
                        c = ws.cell(row=erow, column=PCT_C,
                                    value=f"=ROUND({gt_let}{erow}/600*100,2)")
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        pass_chk = ",".join([f"{sl}{erow}>=35" for sl in sub_lets])
                        c = ws.cell(row=erow, column=RES_C,
                                    value=f'=IF(AND({pass_chk}),"PASS","FAIL")')
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        c.font=Font(name="Arial", bold=True)
                        ws.cell(row=erow, column=REM_C, value="").border = thin
                        c = ws.cell(row=erow, column=RNK_C, value="")
                        c.fill=fl; c.border=thin; c.alignment=ctr
                        c.font=Font(name="Arial", bold=True, color="C00000")
                        h_row = s_idx + 2
                        ws_h.cell(row=h_row, column=1,
                                  value=f"=Consolidated!{gt_let}{erow}")
                        ws_h.cell(row=h_row, column=2,
                                  value=f'=IF(Consolidated!{res_let}{erow}="PASS",1,0)')
                        avg_excel_rows.append((erow, h_row))

                    else:
                        frow = final_df.iloc[s_idx*7 + cat_idx]
                        for i, pc in enumerate(pos_cols):
                            v = frow.get(pc, "")
                            try: v = float(v)
                            except: pass
                            c = ws.cell(row=erow, column=SUB_S+i, value=v)
                            c.fill=fl; c.border=thin; c.alignment=ctr
                        for ri, rc in enumerate(result_cols):
                            v = "" if rc == "Rank" else frow.get(rc, "")
                            c = ws.cell(row=erow, column=GT_C+ri, value=v)
                            c.fill=fl; c.border=thin; c.alignment=ctr

                    for ci in [1, 2, 3]:
                        c = ws.cell(row=erow, column=ci)
                        c.fill=fl; c.border=thin
                        c.font=Font(name="Arial", bold=(ci == 2 and cat_idx == 0))

            for (erow, h_row) in avg_excel_rows:
                rank_formula = (
                    f"=IF(_RankHelper!$B${h_row}=1,"
                    f"COUNTIF({h_gt_rng},\">\"&_RankHelper!$A${h_row})+1,\"\")"
                )
                fl = cat_fill["Average Marks 200/2=100"]
                c  = ws.cell(row=erow, column=RNK_C, value=rank_formula)
                c.fill=fl; c.border=thin; c.alignment=ctr
                c.font=Font(name="Arial", bold=True, color="C00000")

            ws.freeze_panes = "A2"
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.download_button(
                "📥 Download Excel (with Live Formulas)",
                output.getvalue(),
                "Final_Consolidated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── PDF Result Slips — ALWAYS visible once report is generated ─────────
        if st.session_state.get("report_ready"):
            student_results = st.session_state.student_results
            all_students_s  = st.session_state.all_students_snap
            faculty_s       = st.session_state.faculty_snap
            cfg_s           = st.session_state.cfg_snap
            pos_cols_s      = st.session_state.pos_cols_snap
            exam_configs_s  = st.session_state.exam_configs_snap
            sheet_names_s   = st.session_state.xl_sheet_names

            st.markdown("---")
            st.subheader("📄 Generate Exam-wise Result Slip PDFs")

            school_name = st.text_input(
                "🏫 School / College Name",
                value=st.session_state.get("school_name_input", "Your School Name"),
                key="school_name_input",
                help="This appears at the top of every result slip",
            )

            exam_options = [
                ec["label"] for ec in exam_configs_s
                if any(s.strip().upper() in ec["sheets"] for s in sheet_names_s)
            ]
            sel_exams = st.multiselect(
                "Select Exam(s) to generate PDFs for",
                options=exam_options,
                default=exam_options,
                key="sel_exams_pdf",
            )

            if st.button("📄 Generate PDF Result Slips", key="gen_pdf_btn"):
                if not sel_exams:
                    st.warning("Please select at least one exam.")
                else:
                    st.session_state.pdf_results = {}
                    for exam_label in sel_exams:
                        exam_data_for_pdf = {}
                        for sr in student_results:
                            roll   = sr["roll"]
                            subj_6 = sr["subj_6"]
                            ed     = all_students_s[roll]["Exams"].get(exam_label, {})
                            exam_data_for_pdf[roll] = {abbr: ed.get(abbr, "") for abbr in subj_6}
                            exam_data_for_pdf[roll]["Grand Total"] = ed.get("Grand Total", "")
                            exam_data_for_pdf[roll]["%"]           = ed.get("%", "")
                            exam_data_for_pdf[roll]["Result"]      = ed.get("Result", "")

                        pdf_buf = build_exam_pdf(
                            school_name        = school_name,
                            faculty_name       = faculty_s,
                            exam_label         = exam_label,
                            student_results    = student_results,
                            cfg                = cfg_s,
                            pos_cols           = pos_cols_s,
                            selected_exam_data = exam_data_for_pdf,
                        )
                        st.session_state.pdf_results[exam_label] = pdf_buf.getvalue()

            # Show download buttons for all generated PDFs
            if st.session_state.get("pdf_results"):
                st.markdown("#### 📥 Download PDFs")
                cols = st.columns(len(st.session_state.pdf_results))
                for col, (exam_label, pdf_bytes) in zip(
                    cols, st.session_state.pdf_results.items()
                ):
                    safe_name = exam_label.replace("/", "-").replace(" ", "_")
                    n_students = len(student_results)
                    col.download_button(
                        label=f"📥 {exam_label}",
                        data=pdf_bytes,
                        file_name=f"Results_{safe_name}.pdf",
                        mime="application/pdf",
                        key=f"dl_pdf_{safe_name}",
                    )
                    col.caption(f"{n_students} students · {-(-n_students//2)} pages")

    except Exception as e:
        st.error(f"Error: {e}")
        import traceback
        st.code(traceback.format_exc())
